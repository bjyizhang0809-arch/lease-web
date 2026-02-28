"""
生成 public/template.xlsx 模板文件
包含 3 个 Sheet，带格式、说明和示例数据
第 1 行 = 列名（计算器读取时用）
第 2 行 = 字段说明（灰色斜体，用户参考）
第 3-4 行 = 示例数据（可删除）
第 5 行起 = 空白填写区
"""
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'public', 'template.xlsx')

# Colors
BLUE = "DBEAFE"
GRAY = "F5F5F5"
INPUT_BG = "FFFFF0"

def col_header(ws, col, row, value, width=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    c.font = Font(bold=True, size=9)
    c.fill = PatternFill("solid", fgColor=BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width

def col_desc(ws, col, row, value):
    c = ws.cell(row=row, column=col)
    c.value = value
    c.font = Font(size=8, color="777777", italic=True)
    c.fill = PatternFill("solid", fgColor=GRAY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def data_cell(ws, col, row, value):
    c = ws.cell(row=row, column=col)
    c.value = value
    c.font = Font(size=9)
    c.fill = PatternFill("solid", fgColor=GRAY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style='thin', color='CCCCCC')
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def empty_cell(ws, col, row):
    c = ws.cell(row=row, column=col)
    c.fill = PatternFill("solid", fgColor=INPUT_BG)
    thin = Side(style='thin', color='DDDDDD')
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def make_sheet1(wb):
    ws = wb.active
    ws.title = "合同原始数据"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24

    # Row 1: 列名（计算器读取列名用）
    headers = [
        ("客户名称", 20),
        ("商户编号", 12),
        ("交付日", 14),
        ("租期届满日", 16),
        ("免租期", 10),
        ("保底租金第1年（必须）", 22),
        ("保底租金第2年", 16),
        ("保底租金第3年", 16),
        ("保底租金第4年", 16),
        ("保底租金第5年", 16),
        ("保底租金第6年", 16),
        ("保底租金第7年", 16),
    ]
    for i, (header, width) in enumerate(headers):
        col_header(ws, i + 1, 1, header, width)

    # Row 2-3: 示例数据（可删除后填入自己的数据）
    examples = [
        ["北京lbcy餐饮管理有限公司", "B1-01c", date(2025, 5, 12), date(2027, 5, 11), 30, 26496.00, 27820.80, None, None, None, None, None],
        ["上海XX贸易有限公司", "C2-03a", date(2024, 1, 1), date(2028, 12, 31), 60, 120000.00, 126000.00, 132300.00, 138915.00, None, None, None],
    ]
    for r_offset, row_data in enumerate(examples):
        row_num = 2 + r_offset
        ws.row_dimensions[row_num].height = 18
        for c_offset, val in enumerate(row_data):
            c = ws.cell(row=row_num, column=c_offset + 1)
            c.value = val
            c.font = Font(size=9)
            c.fill = PatternFill("solid", fgColor=GRAY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(style='thin', color='CCCCCC')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # Format date columns
            if c_offset in (2, 3) and val is not None:
                c.number_format = 'YYYY-MM-DD'

    # Row 4-23: 空白填写区
    for row_num in range(4, 24):
        ws.row_dimensions[row_num].height = 18
        for col_num in range(1, 13):
            empty_cell(ws, col_num, row_num)


def make_sheet2(wb):
    ws = wb.create_sheet("银行对账单")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24

    headers = [("交易时间", 22), ("贷方发生额（收入）", 20), ("对方户名", 25)]
    for i, (h, w) in enumerate(headers):
        col_header(ws, i + 1, 1, h, w)

    examples = [
        [date(2025, 8, 5), 26496.00, "北京lbcy餐饮管理有限公司"],
        [date(2025, 9, 3), 26496.00, "北京lbcy餐饮管理有限公司"],
    ]
    for r_offset, row_data in enumerate(examples):
        row_num = 2 + r_offset
        ws.row_dimensions[row_num].height = 18
        for c_offset, val in enumerate(row_data):
            c = ws.cell(row=row_num, column=c_offset + 1)
            c.value = val
            c.font = Font(size=9)
            c.fill = PatternFill("solid", fgColor=GRAY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(style='thin', color='CCCCCC')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if c_offset == 0 and val is not None:
                c.number_format = 'YYYY-MM-DD'

    for row_num in range(4, 24):
        ws.row_dimensions[row_num].height = 18
        for col_num in range(1, 4):
            empty_cell(ws, col_num, row_num)


def make_sheet3(wb):
    ws = wb.create_sheet("发票信息汇总表")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24

    headers = [("购买方名称", 25), ("开票日期", 18), ("价税合计", 22)]
    for i, (h, w) in enumerate(headers):
        col_header(ws, i + 1, 1, h, w)

    examples = [
        ["北京lbcy餐饮管理有限公司", date(2025, 8, 10), 29920.48],
        ["北京lbcy餐饮管理有限公司", date(2025, 9, 10), 29920.48],
    ]
    for r_offset, row_data in enumerate(examples):
        row_num = 2 + r_offset
        ws.row_dimensions[row_num].height = 18
        for c_offset, val in enumerate(row_data):
            c = ws.cell(row=row_num, column=c_offset + 1)
            c.value = val
            c.font = Font(size=9)
            c.fill = PatternFill("solid", fgColor=GRAY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(style='thin', color='CCCCCC')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if c_offset == 1 and val is not None:
                c.number_format = 'YYYY-MM-DD'

    for row_num in range(4, 24):
        ws.row_dimensions[row_num].height = 18
        for col_num in range(1, 4):
            empty_cell(ws, col_num, row_num)


def main():
    wb = Workbook()
    make_sheet1(wb)
    make_sheet2(wb)
    make_sheet3(wb)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"✅ 模板已生成: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
